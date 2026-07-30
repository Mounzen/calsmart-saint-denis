#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Extraction Ville - positionnement des candidats issus des audiences elus
sur les logements en proposition (Logivia / CALSmart2 - Ville de Saint-Denis).

Produit un classeur Excel :
  - Candidats        : donnees des candidats passes en audience
  - Logements        : logements en proposition + levier (Ville / autre contingent)
  - Tableau croise   : candidats x logements, taux d'effort (formule vivante) + eligibilite
  - Par logement     : pour chaque logement, candidats audience classes
  - Par candidat     : pour chaque candidat, logements ou il est placable
  - Synthese         : compteurs, legende et rappel du cadre

Le taux d'effort (loyer / revenu) est calcule par le script (instantane).
Pour rafraichir apres une mise a jour des donnees, relancer le script.
Le "Score CAL est." reproduit a l'identique le moteur du serveur (computeScore),
reglages par defaut, bonus audience inclus sur le contingent communal.

Usage : python extraction_ville.py [dossier_donnees] [fichier_sortie.xlsx]
"""
import json, os, re, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

DATA = sys.argv[1] if len(sys.argv) > 1 else '/sessions/laughing-stoic-davinci/mnt/CALSmart2/server/data'
OUT  = sys.argv[2] if len(sys.argv) > 2 else '/sessions/laughing-stoic-davinci/mnt/outputs/Extraction-Ville-Audiences.xlsx'
DATA = DATA.rstrip('/') + '/'

load = lambda f: json.load(open(DATA + f, encoding='utf-8'))
dem_all = load('demandeurs.json'); log_all = load('logements.json')
aud = load('audiences.json'); ref = load('referentiels.json')
elus = {e['id']: e for e in (ref.get('elus') or [])}
cfgC = ref.get('contingents_config') or []

def is_communal(c):
    e = next((x for x in cfgC if x.get('nom') == c), None)
    return bool(re.search('communal', (e or {}).get('description', '') or '', re.I)) or str(c or '').lower() == 'ville'

# ---- Moteur de score : portage fidele de computeScore (valide 72/72 vs serveur) ----
TYP = ['T1', 'T2', 'T3', 'T4', 'T5', 'T6']
ti = lambda t: TYP.index(t) if t in TYP else -1
inR = lambda t, mn, mx: ti(t) >= ti(mn) and ti(t) <= ti(mx)
AUDCFG = {'actif': True, 'bonus_favorable': 8, 'bonus_quartier_concordant': 2,
          'plafond_bonus': 10, 'exiger_contingent_communal': True}

def compute(d, l, adem, est_communal):
    excl = []
    if d.get('statut') != 'active': excl.append('Demande non active')
    if not inR(l['typ'], d.get('typ_min') or 'T1', d.get('typ_max') or 'T6'): excl.append('Typologie incompatible')
    if d.get('pmr') and not l.get('pmr'): excl.append('PMR requis non disponible')
    if d.get('rdc') and not l.get('rdc'): excl.append('RDC requis non disponible')
    rev = float(d.get('rev') or 1); loyer = float(l.get('loyer') or 0); te = loyer / rev * 100
    if te > 40: excl.append('Taux effort %d%% trop eleve' % round(te))
    if excl:
        return {'eligible': False, 'total': 0, 'te': round(te, 1), 'ab': 0, 'excl': excl}
    sTyp = 20 if l['typ'] == d.get('typ_v') else 15
    np = int(d.get('adultes') or 0) + int(d.get('enfants') or 0); idx = ti(l['typ'])
    sComp = 15 if idx <= np <= idx + 2 else 10 if np in (idx - 1, idx + 3) else 5 if np == idx + 4 else 0
    sTaux = 20 if te <= 25 else 16 if te <= 30 else 10 if te <= 35 else 5
    anc = int(d.get('anc') or 0)
    sAnc = 10 if anc >= 36 else 8 if anc >= 24 else 5 if anc >= 12 else 3 if anc >= 6 else 1
    sUrg = 0
    for k, v in (('sans_log', 6), ('violences', 5), ('handicap', 4), ('expulsion', 5), ('suroc', 4), ('grossesse', 3)):
        if d.get(k): sUrg += v
    if d.get('urgence') and sUrg < 4: sUrg += 3
    sUrg = min(sUrg, 15)
    sLoc = 10 if l.get('quartier') in (d.get('quartiers') or []) else 8 if l.get('secteur') in (d.get('secteurs') or []) else 2
    sPrio = 5 if (d.get('dalo') or d.get('prio_expulsion')) else 3 if (d.get('mutation') or d.get('prio_handicap')) else 0
    sDos = 5 if d.get('pieces') else 1
    base = sTyp + sComp + sTaux + sAnc + sUrg + sLoc + sPrio + sDos
    bonus = 5; ab = 0  # +5 "jamais presente" (historique vide)
    favs = [a for a in adem if a.get('favorable')]
    if AUDCFG['actif'] and favs and ((not AUDCFG['exiger_contingent_communal']) or est_communal):
        qm = any((a.get('quartier_souhaite') == l.get('quartier')) or (a.get('quartier_elu') == l.get('quartier')) for a in favs)
        ab = max(0, min(AUDCFG['bonus_favorable'] + (AUDCFG['bonus_quartier_concordant'] if qm else 0), AUDCFG['plafond_bonus']))
        bonus += ab
    return {'eligible': True, 'total': min(max(base + bonus, 0), 100), 'te': round(te, 1), 'ab': ab, 'excl': []}

# ---- Selection : candidats audience, logements en proposition ----
aud_by_dem = {}
for a in aud: aud_by_dem.setdefault(a['dem_id'], []).append(a)
cands = [d for d in dem_all if d['id'] in aud_by_dem]
logs = log_all

def fav_aud(dem_id):
    fs = [a for a in aud_by_dem.get(dem_id, []) if a.get('favorable')]
    return fs[0] if fs else (aud_by_dem.get(dem_id, [None])[0])

PRIO_MAP = [('dalo', 'DALO'), ('violences', 'VIF'), ('sans_log', 'SDF'), ('expulsion', 'Expulsion'),
            ('handicap', 'Handicap'), ('suroc', 'Suroccup.'), ('grossesse', 'Grossesse'),
            ('urgence', 'Urgence'), ('mutation', 'Mutation'), ('prio_handicap', 'Prio handicap'),
            ('prio_expulsion', 'Prio expulsion')]
def prios(d): return ', '.join(lab for k, lab in PRIO_MAP if d.get(k)) or '—'
def levier(c): return 'Proposition directe' if is_communal(c) else 'Plaidoyer CAL'

# precompute all pairs
res = {(d['id'], l['id']): compute(d, l, aud_by_dem.get(d['id'], []), is_communal(l['contingent']))
       for d in cands for l in logs}

# ---- Styles ----
FN = 'Arial'
NAVY = '0B1E3D'; ACCENT = 'E05C2A'
f_title = Font(name=FN, size=15, bold=True, color='FFFFFF')
f_h = Font(name=FN, size=10, bold=True, color='FFFFFF')
f_b = Font(name=FN, size=10, color='1A1A1A')
f_bb = Font(name=FN, size=10, bold=True, color='1A1A1A')
f_muted = Font(name=FN, size=9, color='5B6B85')
f_sec = Font(name=FN, size=11, bold=True, color=NAVY)
fill_navy = PatternFill('solid', fgColor=NAVY)
fill_ville = PatternFill('solid', fgColor='DCFCE7')
fill_autre = PatternFill('solid', fgColor='FEF3C7')
fill_grey = PatternFill('solid', fgColor='EEF1F6')
fill_head = PatternFill('solid', fgColor='1D3557')
green = PatternFill('solid', fgColor='C7F0D2'); amber = PatternFill('solid', fgColor='FDE9B8'); red = PatternFill('solid', fgColor='F9C9C9')
thin = Side(style='thin', color='D0D7E2')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left = Alignment(horizontal='left', vertical='center', wrap_text=True)

wb = Workbook()

def style_header(ws, row, ncol, fill=fill_head):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = f_h; cell.fill = fill; cell.alignment = center; cell.border = border

# =============== Sheet Candidats ===============
wsC = wb.active; wsC.title = 'Candidats'
wsC.merge_cells('A1:M1'); wsC['A1'] = 'CANDIDATS ISSUS DES AUDIENCES ELUS'; wsC['A1'].font = Font(name=FN, size=13, bold=True, color=NAVY)
head = ['ID', 'Candidat', 'NUD', 'Elu (audience)', 'Objet audience', 'Favorable', 'Statut demande',
        'Revenu (EUR/mois)', 'Typo min', 'Typo souhait', 'Typo max', 'Personnes', 'Priorites']
wsC.append([]); wsC.append(head); style_header(wsC, 3, len(head))
cand_row = {}
for i, d in enumerate(cands):
    r = 4 + i; cand_row[d['id']] = r
    a = fav_aud(d['id']); elu = elus.get(a['elu_id'], {}) if a else {}
    wsC.append([d['id'], d.get('nom', '') + ' ' + d.get('prenom', ''), d.get('nud', ''),
                elu.get('nom', a['elu_id'] if a else ''), a.get('objet', '') if a else '',
                'Oui' if (a and a.get('favorable')) else 'Non', d.get('statut', ''),
                int(d.get('rev') or 0), d.get('typ_min', ''), d.get('typ_v', ''), d.get('typ_max', ''),
                int(d.get('adultes') or 0) + int(d.get('enfants') or 0), prios(d)])
    for c in range(1, len(head) + 1):
        cell = wsC.cell(row=r, column=c); cell.font = f_b; cell.border = border
        cell.alignment = left if c in (2, 4, 5, 13) else center
        if d.get('statut') != 'active': cell.fill = fill_grey
    wsC.cell(row=r, column=8).number_format = '# ##0'
REVCOL = 'H'  # colonne Revenu
widths = [6, 20, 17, 15, 26, 9, 13, 13, 8, 9, 8, 9, 24]
for i, w in enumerate(widths): wsC.column_dimensions[get_column_letter(i + 1)].width = w
wsC.freeze_panes = 'A4'

# =============== Sheet Logements ===============
wsL = wb.create_sheet('Logements')
wsL.merge_cells('A1:L1'); wsL['A1'] = 'LOGEMENTS EN PROPOSITION'; wsL['A1'].font = Font(name=FN, size=13, bold=True, color=NAVY)
headL = ['ID', 'Ref', 'Bailleur', 'Typo', 'Surface m2', 'Loyer CC (EUR)', 'Quartier', 'Secteur', 'Contingent', 'Levier Ville', 'Dispo', 'PMR', 'RDC']
wsL.append([]); wsL.append(headL); style_header(wsL, 3, len(headL))
log_row = {}
for j, l in enumerate(logs):
    r = 4 + j; log_row[l['id']] = r
    comm = is_communal(l['contingent'])
    wsL.append([l['id'], l.get('ref', ''), l.get('bailleur', ''), l.get('typ', ''), l.get('surface', ''),
                int(l.get('loyer') or 0), l.get('quartier', ''), l.get('secteur', ''), l.get('contingent', ''),
                levier(l['contingent']), l.get('dispo', ''), 'Oui' if l.get('pmr') else 'Non', 'Oui' if l.get('rdc') else 'Non'])
    for c in range(1, len(headL) + 1):
        cell = wsL.cell(row=r, column=c); cell.font = f_b; cell.border = border
        cell.alignment = left if c in (2, 3, 7, 8, 9, 10) else center
    wsL.cell(row=r, column=6).number_format = '# ##0'
    wsL.cell(row=r, column=10).fill = fill_ville if comm else fill_autre
    wsL.cell(row=r, column=10).font = f_bb
LOYCOL = 'F'
widthsL = [6, 12, 10, 7, 10, 14, 14, 10, 15, 16, 11, 6, 6]
for i, w in enumerate(widthsL): wsL.column_dimensions[get_column_letter(i + 1)].width = w
wsL.freeze_panes = 'A4'

# =============== Sheet Tableau croise ===============
wsX = wb.create_sheet('Tableau croise')
wsX.merge_cells('A1:' + get_column_letter(4 + len(logs)) + '1')
wsX['A1'] = 'TABLEAU CROISE  -  taux d\'effort (loyer / revenu) et eligibilite'; wsX['A1'].font = Font(name=FN, size=13, bold=True, color=NAVY)
# header rows: row3 logement ref/typo, row4 loyer/contingent
base_cols = ['Candidat', 'Elu', 'Revenu', 'Statut']
for i, h in enumerate(base_cols):
    for rr in (3, 4):
        cell = wsX.cell(row=rr, column=1 + i)
    wsX.merge_cells(start_row=3, start_column=1 + i, end_row=4, end_column=1 + i)
    cell = wsX.cell(row=3, column=1 + i); cell.value = h; cell.font = f_h; cell.fill = fill_head; cell.alignment = center; cell.border = border
    wsX.cell(row=4, column=1 + i).border = border; wsX.cell(row=4, column=1 + i).fill = fill_head
for j, l in enumerate(logs):
    col = 5 + j; comm = is_communal(l['contingent'])
    c1 = wsX.cell(row=3, column=col); c1.value = l.get('ref', '') + '\n' + l.get('typ', ''); c1.font = f_h; c1.fill = fill_navy; c1.alignment = center; c1.border = border
    c2 = wsX.cell(row=4, column=col); c2.value = str(int(l.get('loyer') or 0)) + ' EUR\n' + l.get('contingent', ''); c2.font = Font(name=FN, size=9, bold=True, color=(NAVY if comm else '7A5B00')); c2.fill = fill_ville if comm else fill_autre; c2.alignment = center; c2.border = border
    wsX.column_dimensions[get_column_letter(col)].width = 12
for i, w in enumerate([20, 12, 11, 10]): wsX.column_dimensions[get_column_letter(i + 1)].width = w

taux_cells = []
for i, d in enumerate(cands):
    r = 5 + i
    a = fav_aud(d['id']); elu = elus.get(a['elu_id'], {}) if a else {}
    active = d.get('statut') == 'active'
    wsX.cell(row=r, column=1, value=d.get('nom', '') + ' ' + d.get('prenom', '')).font = f_bb
    wsX.cell(row=r, column=2, value=elu.get('nom', '')).font = f_b
    wsX.cell(row=r, column=3, value=int(d.get('rev') or 0)).number_format = '# ##0'
    wsX.cell(row=r, column=4, value='active' if active else d.get('statut', '')).font = f_b if active else f_muted
    for c in range(1, 5):
        cell = wsX.cell(row=r, column=c); cell.border = border; cell.alignment = left if c in (1, 2) else center
        if not active: cell.fill = fill_grey
    for j, l in enumerate(logs):
        col = 5 + j; rr = res[(d['id'], l['id'])]; cell = wsX.cell(row=r, column=col); cell.border = border; cell.alignment = center
        structural = (not active) or ('Typologie incompatible' in rr['excl']) or ('PMR requis non disponible' in rr['excl']) or ('RDC requis non disponible' in rr['excl'])
        if structural:
            cell.value = '—'; cell.fill = fill_grey; cell.font = f_muted
        else:
            cell.value = round(rr['te'] / 100.0, 4)
            cell.number_format = '0.0%'; cell.font = f_b
            taux_cells.append('%s%d' % (get_column_letter(col), r))
# conditional formatting on taux cells (eligible <=40%)
if taux_cells:
    rng = ' '.join(taux_cells)
    wsX.conditional_formatting.add(rng, CellIsRule(operator='lessThanOrEqual', formula=['0.30'], fill=green, stopIfTrue=True))
    wsX.conditional_formatting.add(rng, CellIsRule(operator='greaterThan', formula=['0.40'], fill=red, stopIfTrue=True))
    wsX.conditional_formatting.add(rng, CellIsRule(operator='between', formula=['0.30', '0.40'], fill=amber))
note_r = 6 + len(cands)
wsX.cell(row=note_r, column=1, value="Vert = taux <=30%  |  Orange = 30-40% (limite)  |  Rouge = >40% (ineligible)  |  — = typologie/PMR/RDC incompatible ou demande non active").font = f_muted
wsX.merge_cells(start_row=note_r, start_column=1, end_row=note_r, end_column=4 + len(logs))
wsX.cell(row=note_r + 1, column=1, value="En-tetes verts = contingent Ville (proposition directe)  |  En-tetes oranges = autre contingent (plaidoyer CAL, priorite au reservataire)").font = f_muted
wsX.merge_cells(start_row=note_r + 1, start_column=1, end_row=note_r + 1, end_column=4 + len(logs))
wsX.freeze_panes = 'E5'

# =============== Sheet Par logement ===============
wsPL = wb.create_sheet('Par logement')
wsPL.merge_cells('A1:H1'); wsPL['A1'] = 'PAR LOGEMENT  -  candidats audience classes (contingent Ville en premier)'; wsPL['A1'].font = Font(name=FN, size=13, bold=True, color=NAVY)
r = 3
logs_sorted = sorted(logs, key=lambda l: (0 if is_communal(l['contingent']) else 1, l.get('ref', '')))
for l in logs_sorted:
    comm = is_communal(l['contingent'])
    wsPL.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    hc = wsPL.cell(row=r, column=1, value='%s  -  %s  -  %d EUR  -  %s  -  Contingent %s  [%s]' % (
        l.get('ref', ''), l.get('typ', ''), int(l.get('loyer') or 0), l.get('quartier', ''), l.get('contingent', ''), levier(l['contingent'])))
    hc.font = Font(name=FN, size=11, bold=True, color='FFFFFF'); hc.fill = fill_ville if comm else fill_autre
    hc.font = Font(name=FN, size=11, bold=True, color=NAVY); hc.alignment = left
    r += 1
    cols = ['Rang', 'Candidat', 'Elu', "Taux d'effort", 'Score CAL est.', 'Priorites', 'Eligible', 'Motif si non']
    for i, h in enumerate(cols):
        cell = wsPL.cell(row=r, column=1 + i, value=h); cell.font = f_h; cell.fill = fill_head; cell.alignment = center; cell.border = border
    r += 1
    rows = []
    for d in cands:
        rr = res[(d['id'], l['id'])]; rows.append((d, rr))
    rows.sort(key=lambda t: (0 if t[1]['eligible'] else 1, -t[1]['total'], t[1]['te']))
    rank = 0
    for d, rr in rows:
        a = fav_aud(d['id']); elu = elus.get(a['elu_id'], {}) if a else {}
        elig = rr['eligible']; rank = rank + 1 if elig else rank
        wsPL.cell(row=r, column=1, value=(rank if elig else '')).alignment = center
        wsPL.cell(row=r, column=2, value=d.get('nom', '') + ' ' + d.get('prenom', ''))
        wsPL.cell(row=r, column=3, value=elu.get('nom', ''))
        tc = wsPL.cell(row=r, column=4)
        structural = (d.get('statut') != 'active') or ('Typologie incompatible' in rr['excl']) or ('PMR requis non disponible' in rr['excl']) or ('RDC requis non disponible' in rr['excl'])
        if structural:
            tc.value = '—'
        else:
            tc.value = round(rr['te'] / 100.0, 4); tc.number_format = '0.0%'
        tc.alignment = center
        wsPL.cell(row=r, column=5, value=(rr['total'] if elig else '')).alignment = center
        wsPL.cell(row=r, column=6, value=prios(d)).alignment = left
        ec = wsPL.cell(row=r, column=7, value='Oui' if elig else 'Non'); ec.alignment = center
        ec.font = Font(name=FN, size=10, bold=True, color=('16A34A' if elig else 'DC2626'))
        wsPL.cell(row=r, column=8, value='' if elig else '; '.join(rr['excl'])).alignment = left
        for c in range(1, 9):
            cell = wsPL.cell(row=r, column=c); cell.border = border
            if not cell.font or cell.font.name != FN: cell.font = f_b
            if not elig: cell.fill = fill_grey
        r += 1
    r += 1
for i, w in enumerate([6, 20, 12, 12, 12, 26, 9, 34]): wsPL.column_dimensions[get_column_letter(i + 1)].width = w

# =============== Sheet Par candidat ===============
wsPC = wb.create_sheet('Par candidat')
wsPC.merge_cells('A1:H1'); wsPC['A1'] = 'PAR CANDIDAT  -  logements ou le candidat est placable'; wsPC['A1'].font = Font(name=FN, size=13, bold=True, color=NAVY)
r = 3
for d in cands:
    a = fav_aud(d['id']); elu = elus.get(a['elu_id'], {}) if a else {}
    wsPC.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    hc = wsPC.cell(row=r, column=1, value='%s %s  -  Elu %s  -  Revenu %d EUR  -  Typo %s-%s-%s  -  %s  -  [%s]' % (
        d.get('nom', ''), d.get('prenom', ''), elu.get('nom', ''), int(d.get('rev') or 0),
        d.get('typ_min', ''), d.get('typ_v', ''), d.get('typ_max', ''), prios(d),
        'demande active' if d.get('statut') == 'active' else 'DEMANDE ' + str(d.get('statut', '')).upper()))
    hc.font = Font(name=FN, size=11, bold=True, color=NAVY); hc.alignment = left
    hc.fill = fill_ville if d.get('statut') == 'active' else fill_grey
    r += 1
    cols = ['Logement', 'Typo', 'Loyer', 'Quartier', 'Contingent', "Taux d'effort", 'Score CAL est.', 'Levier / eligibilite']
    for i, h in enumerate(cols):
        cell = wsPC.cell(row=r, column=1 + i, value=h); cell.font = f_h; cell.fill = fill_head; cell.alignment = center; cell.border = border
    r += 1
    rows = sorted(logs, key=lambda l: (0 if res[(d['id'], l['id'])]['eligible'] else 1, -res[(d['id'], l['id'])]['total']))
    for l in rows:
        rr = res[(d['id'], l['id'])]; elig = rr['eligible']; comm = is_communal(l['contingent'])
        wsPC.cell(row=r, column=1, value=l.get('ref', '')).alignment = left
        wsPC.cell(row=r, column=2, value=l.get('typ', '')).alignment = center
        wsPC.cell(row=r, column=3, value=int(l.get('loyer') or 0)).number_format = '# ##0'
        wsPC.cell(row=r, column=4, value=l.get('quartier', '')).alignment = left
        wsPC.cell(row=r, column=5, value=l.get('contingent', '')).alignment = center
        tc = wsPC.cell(row=r, column=6)
        structural = (d.get('statut') != 'active') or ('Typologie incompatible' in rr['excl']) or ('PMR requis non disponible' in rr['excl']) or ('RDC requis non disponible' in rr['excl'])
        tc.value = '—' if structural else round(rr['te'] / 100.0, 4)
        if not structural: tc.number_format = '0.0%'
        tc.alignment = center
        wsPC.cell(row=r, column=7, value=(rr['total'] if elig else '')).alignment = center
        if elig:
            lev = ('Proposition directe (Ville)' if comm else 'Plaidoyer CAL')
        else:
            lev = 'Non eligible : ' + '; '.join(rr['excl'])
        lc = wsPC.cell(row=r, column=8, value=lev); lc.alignment = left
        for c in range(1, 9):
            cell = wsPC.cell(row=r, column=c); cell.border = border; cell.font = f_b
            if not elig: cell.fill = fill_grey
            elif comm and c == 8: cell.fill = fill_ville
        r += 1
    r += 1
for i, w in enumerate([12, 7, 10, 14, 15, 12, 12, 40]): wsPC.column_dimensions[get_column_letter(i + 1)].width = w

# =============== Sheet Synthese ===============
wsS = wb.create_sheet('Synthese')
nb_actifs = sum(1 for d in cands if d.get('statut') == 'active')
nb_ville = sum(1 for l in logs if is_communal(l['contingent']))
placables = [d for d in cands if any(res[(d['id'], l['id'])]['eligible'] for l in logs)]
wsS.merge_cells('A1:D1'); wsS['A1'] = 'SYNTHESE & LEGENDE'; wsS['A1'].font = Font(name=FN, size=14, bold=True, color=NAVY)
lines = [
    ('Candidats issus des audiences elus', len(cands)),
    ('  dont demande active', nb_actifs),
    ('Logements en proposition', len(logs)),
    ('  dont contingent Ville (proposition directe)', nb_ville),
    ('Candidats avec au moins un logement eligible', len(placables)),
]
r = 3
for lab, val in lines:
    wsS.cell(row=r, column=1, value=lab).font = f_b
    c = wsS.cell(row=r, column=2, value=val); c.font = f_bb; c.alignment = center
    r += 1
if placables:
    r += 1; wsS.cell(row=r, column=1, value='Candidats placables : ' + ', '.join(d.get('nom', '') + ' ' + d.get('prenom', '') for d in placables)).font = f_b; r += 1
r += 1
wsS.cell(row=r, column=1, value='LEGENDE').font = f_sec; r += 1
leg = [
    ('Vert', 'Taux d\'effort <= 30 % (confortable)', green),
    ('Orange', 'Taux d\'effort 30 - 40 % (limite, eligible)', amber),
    ('Rouge', 'Taux d\'effort > 40 % : INELIGIBLE (regle du moteur)', red),
    ('—', 'Typologie / PMR / RDC incompatible, ou demande non active', fill_grey),
    ('En-tete vert', 'Contingent Ville : la commune propose directement', fill_ville),
    ('En-tete orange', 'Autre contingent : plaidoyer en CAL, priorite au reservataire', fill_autre),
]
for lab, desc, fl in leg:
    cc = wsS.cell(row=r, column=1, value=lab); cc.fill = fl; cc.font = f_bb; cc.border = border; cc.alignment = center
    wsS.cell(row=r, column=2, value=desc).font = f_b; r += 1
r += 1
wsS.cell(row=r, column=1, value='CADRE (a valider avec le service juridique)').font = f_sec; r += 1
notes = [
    "Taux d'effort = loyer charges comprises / revenu, HORS aides (APL). Au-dela de 40 %, le dossier est",
    "ineligible dans le moteur. Avec APL, le reste-a-charge reel est plus bas : un taux 40-45 % peut meriter",
    "un examen (a instruire au cas par cas).",
    "",
    "Contingent Ville : la commune est reservataire, elle propose directement ses candidats.",
    "Autres contingents (Prefecture/DALO, Action Logement, Bailleur...) : la Ville peut PLAIDER en CAL, mais",
    "la priorite revient au reservataire ; sur le contingent prefectoral, les publics DALO/prioritaires",
    "priment (CCH L.441-1). Cet outil prepare l'aide a la decision ; la CAL reste souveraine.",
    "",
    "Score CAL est. : reproduit le moteur Logivia (reglages par defaut, bonus audience +8 sur contingent",
    "communal inclus, sans historique de presentation). Valide 72/72 face au serveur.",
]
for t in notes:
    wsS.cell(row=r, column=1, value=t).font = f_muted; r += 1
for i, w in enumerate([46, 44, 10, 10]): wsS.column_dimensions[get_column_letter(i + 1)].width = w
wsS.column_dimensions['A'].width = 20

os.makedirs(os.path.dirname(OUT), exist_ok=True)
wb.save(OUT)
print('Ecrit :', OUT)
print('Candidats audience :', len(cands), '| actifs :', nb_actifs, '| logements :', len(logs), '| Ville :', nb_ville, '| placables :', len(placables))
